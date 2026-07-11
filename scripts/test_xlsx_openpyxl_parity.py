#!/usr/bin/env python3
"""Tests for the manifest-aware `.xlsx` openpyxl oracle harness."""

from __future__ import annotations

import datetime
import importlib.util
import json
import os
from pathlib import Path
import subprocess
import sys
import tempfile
import textwrap
import unittest
from zipfile import ZIP_DEFLATED, ZipFile

try:
    import openpyxl
except ImportError:  # pragma: no cover - local verification provides a venv.
    openpyxl = None


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "xlsx-openpyxl-parity.py"


@unittest.skipIf(openpyxl is None, "openpyxl is not installed")
class XlsxOpenpyxlParityTests(unittest.TestCase):
    def test_skip_classification_maps_rows_to_decisions_and_evidence(self) -> None:
        module = _load_xlsx_openpyxl_parity()

        cases = [
            (
                "openpyxl-unreadable",
                "BadZipFile: Bad magic number for central directory",
                ("excluded_malformed_container", "openpyxl_bad_zip", None),
            ),
            (
                "openpyxl-unreadable",
                "BadZipFile: File is not a zip file",
                ("needs_corpus_crosscheck", "openpyxl_non_zip_container", None),
            ),
            (
                "openpyxl-unreadable",
                "UnsupportedOpenpyxlInlineStringValue: openpyxl silently drops malformed inline-string cells",
                ("documented_oracle_limitation", "openpyxl_malformed_inline_string", None),
            ),
            (
                "openpyxl-unreadable",
                "UnsupportedOpenpyxlNamespace: openpyxl silently drops cells from legacy beta SpreadsheetML namespace",
                ("documented_oracle_limitation", "openpyxl_namespace_limitation", None),
            ),
            (
                "openpyxl-unreadable",
                "UnsupportedOpenpyxlSharedStringsPathVariant: openpyxl requires xl/sharedStrings.xml but package uses xl/SharedStrings.xml",
                (
                    "documented_oracle_limitation",
                    "openpyxl_shared_strings_path_variant",
                    None,
                ),
            ),
            (
                "openpyxl-unreadable",
                "UnsupportedOpenpyxlEmptyStyleIndex: openpyxl treats empty cell style indexes as list keys",
                ("documented_oracle_limitation", "openpyxl_empty_style_index", None),
            ),
            (
                "openpyxl-unreadable",
                "TypeError: expected <class 'openpyxl.styles.fills.Fill'>",
                (
                    "documented_oracle_limitation",
                    "openpyxl_style_parser_limitation",
                    None,
                ),
            ),
            (
                "openpyxl-unreadable",
                "TypeError: CellStyle.__init__() got an unexpected keyword argument 'ZMIENNA_contentFontsCount'",
                (
                    "documented_oracle_limitation",
                    "openpyxl_style_parser_limitation",
                    None,
                ),
            ),
            (
                "openpyxl-unreadable",
                "TypeError: <class 'openpyxl.styles.named_styles._NamedCellStyle'>.name should be <class 'str'> but value is <class 'NoneType'>",
                (
                    "documented_oracle_limitation",
                    "openpyxl_style_parser_limitation",
                    None,
                ),
            ),
            (
                "openpyxl-unreadable",
                "UnsupportedOpenpyxlRootStylesPathVariant: openpyxl requires xl/styles.xml but package uses styles.xml",
                (
                    "documented_oracle_limitation",
                    "openpyxl_root_styles_path_variant",
                    None,
                ),
            ),
            (
                "openpyxl-unreadable",
                "TypeError: Nested.from_tree() missing 1 required positional argument: 'node'",
                (
                    "documented_oracle_limitation",
                    "openpyxl_pivot_parser_limitation",
                    None,
                ),
            ),
            (
                "oversized-comparison",
                "combined text length 255099 exceeds 200000",
                ("needs_bounded_oracle", "comparison_budget_exceeded", None),
            ),
            (
                "oversized-worksheet",
                "Sheet1: declared cells 17179869184 exceed 5000000",
                ("needs_bounded_oracle", "worksheet_guard_exceeded", None),
            ),
            (
                "bounded-shared-string-expansion",
                "expanded shared-string text 80 exceeds 50 hash budget",
                (
                    "documented_bounded_extraction",
                    "shared_string_expansion_budget_exceeded",
                    None,
                ),
            ),
        ]

        for kind, reason, expected in cases:
            with self.subTest(kind=kind, reason=reason):
                self.assertEqual(module.skip_classification(kind, reason), expected)

    def test_corpus_report_crosscheck_refines_non_zip_skip_classification(self) -> None:
        module = _load_xlsx_openpyxl_parity()

        with tempfile.TemporaryDirectory() as tmp:
            corpus_report = Path(tmp) / "corpus-report.txt"
            corpus_report.write_text(
                "failure: .xlsx test-data/spreadsheet/protected.xlsx "
                "kind=unsupported_encrypted_ooxml decision=unsupported_encrypted "
                "evidence=encrypted_ooxml_package container=ole2 "
                "extension_mismatch=true parse: unsupported encrypted OOXML package\n",
                encoding="utf-8",
            )

            failures = module.parse_corpus_report(corpus_report)
            full_path = (
                Path(tmp)
                / "local"
                / "public-corpus"
                / "apache-poi"
                / "test-data"
                / "spreadsheet"
                / "protected.xlsx"
            )

            self.assertEqual(
                module.skip_classification(
                    "openpyxl-unreadable",
                    "BadZipFile: File is not a zip file",
                    path=full_path,
                    corpus_failures=failures,
                ),
                (
                    "unsupported_encrypted",
                    "encrypted_ooxml_package",
                    "unsupported_encrypted_ooxml",
                ),
            )

    def test_shared_strings_path_variants_are_documented_oracle_limitations(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            normal_path = base / "normal.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.title = "Normal"
            workbook.active["A1"] = "ok"
            workbook.save(normal_path)

            variant_path = base / "shared-strings-case.xlsx"
            _write_shared_strings_path_variant_workbook(variant_path, "xl/SharedStrings.xml")

            manifest_path = base / "manifest.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "source": "unit",
                                "path": "normal.xlsx",
                                "local_path": str(normal_path),
                                "status": "downloaded",
                            },
                            {
                                "source": "unit",
                                "path": "shared-strings-case.xlsx",
                                "local_path": str(variant_path),
                                "status": "downloaded",
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )

            fake_extract = base / "fake-rxls-extract"
            fake_extract.write_text(
                textwrap.dedent(
                    """\
                    #!/usr/bin/env python3
                    import pathlib
                    if pathlib.Path(__import__("sys").argv[1]).name == "normal.xlsx":
                        print("# Normal")
                        print("ok")
                    else:
                        print("# Sheet1")
                        print("hello")
                    """
                ),
                encoding="utf-8",
            )
            os.chmod(fake_extract, 0o755)

            output = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT),
                    "--manifest",
                    str(manifest_path),
                    "--bin",
                    str(fake_extract),
                    "--show-skips",
                    "5",
                    "--min",
                    "0.0",
                ],
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn("files: 2", output.stdout)
            self.assertIn("rxls extracted: 2", output.stdout)
            self.assertIn("openpyxl-unreadable: 1", output.stdout)
            self.assertIn("comparable: 1", output.stdout)
            self.assertIn(
                "by_skip_decision: documented_oracle_limitation skipped=1",
                output.stdout,
            )
            self.assertIn(
                "by_skip_evidence: openpyxl_shared_strings_path_variant skipped=1",
                output.stdout,
            )
            self.assertIn(
                "skip: kind=openpyxl-unreadable decision=documented_oracle_limitation "
                "evidence=openpyxl_shared_strings_path_variant",
                output.stdout,
            )

    def test_missing_content_types_is_not_reclassified_by_shared_strings_variant(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            normal_path = base / "normal.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.title = "Normal"
            workbook.active["A1"] = "ok"
            workbook.save(normal_path)

            malformed_path = base / "missing-content-types.xlsx"
            _write_shared_strings_path_variant_workbook(
                malformed_path,
                "xl\\sharedStrings.xml",
                include_content_types=False,
            )

            manifest_path = base / "manifest.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "source": "unit",
                                "path": "normal.xlsx",
                                "local_path": str(normal_path),
                                "status": "downloaded",
                            },
                            {
                                "source": "unit",
                                "path": "missing-content-types.xlsx",
                                "local_path": str(malformed_path),
                                "status": "downloaded",
                            },
                        ]
                    }
                ),
                encoding="utf-8",
            )

            fake_extract = base / "fake-rxls-extract"
            fake_extract.write_text(
                textwrap.dedent(
                    """\
                    #!/usr/bin/env python3
                    import pathlib
                    if pathlib.Path(__import__("sys").argv[1]).name == "normal.xlsx":
                        print("# Normal")
                        print("ok")
                    else:
                        print("# Sheet1")
                        print("hello")
                    """
                ),
                encoding="utf-8",
            )
            os.chmod(fake_extract, 0o755)

            output = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT),
                    "--manifest",
                    str(manifest_path),
                    "--bin",
                    str(fake_extract),
                    "--show-skips",
                    "5",
                    "--min",
                    "0.0",
                ],
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn("files: 2", output.stdout)
            self.assertIn("comparable: 1", output.stdout)
            self.assertIn(
                "by_skip_decision: excluded_malformed_container skipped=1",
                output.stdout,
            )
            self.assertIn(
                "by_skip_evidence: openpyxl_missing_content_types skipped=1",
                output.stdout,
            )

    def test_empty_style_indexes_are_documented_oracle_limitations(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            normal_path = base / "normal.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.title = "Normal"
            workbook.active["A1"] = "ok"
            workbook.save(normal_path)

            empty_style_path = base / "empty-style.xlsx"
            _write_empty_style_index_workbook(empty_style_path)

            manifest_path = base / "manifest.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "source": "unit",
                                "path": "normal.xlsx",
                                "local_path": str(normal_path),
                                "status": "downloaded",
                            },
                            {
                                "source": "unit",
                                "path": "empty-style.xlsx",
                                "local_path": str(empty_style_path),
                                "status": "downloaded",
                            },
                        ]
                    }
                ),
                encoding="utf-8",
            )

            fake_extract = base / "fake-rxls-extract"
            fake_extract.write_text(
                textwrap.dedent(
                    """\
                    #!/usr/bin/env python3
                    import pathlib
                    if pathlib.Path(__import__("sys").argv[1]).name == "normal.xlsx":
                        print("# Normal")
                        print("ok")
                    else:
                        print("# Sheet1")
                        print("empty")
                    """
                ),
                encoding="utf-8",
            )
            os.chmod(fake_extract, 0o755)

            output = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT),
                    "--manifest",
                    str(manifest_path),
                    "--bin",
                    str(fake_extract),
                    "--show-skips",
                    "5",
                    "--min",
                    "0.0",
                ],
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn("files: 2", output.stdout)
            self.assertIn("rxls extracted: 2", output.stdout)
            self.assertIn("openpyxl-unreadable: 1", output.stdout)
            self.assertIn("comparable: 1", output.stdout)
            self.assertIn(
                "by_skip_decision: documented_oracle_limitation skipped=1",
                output.stdout,
            )
            self.assertIn(
                "by_skip_evidence: openpyxl_empty_style_index skipped=1",
                output.stdout,
            )

    def test_root_styles_path_variants_are_documented_oracle_limitations(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            normal_path = base / "normal.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.title = "Normal"
            workbook.active["A1"] = "ok"
            workbook.save(normal_path)

            root_styles_path = base / "root-styles.xlsx"
            _write_root_styles_path_variant_workbook(root_styles_path)

            manifest_path = base / "manifest.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "source": "unit",
                                "path": "normal.xlsx",
                                "local_path": str(normal_path),
                                "status": "downloaded",
                            },
                            {
                                "source": "unit",
                                "path": "root-styles.xlsx",
                                "local_path": str(root_styles_path),
                                "status": "downloaded",
                            },
                        ]
                    }
                ),
                encoding="utf-8",
            )

            fake_extract = base / "fake-rxls-extract"
            fake_extract.write_text(
                textwrap.dedent(
                    """\
                    #!/usr/bin/env python3
                    import pathlib
                    name = pathlib.Path(__import__("sys").argv[1]).name
                    if name == "normal.xlsx":
                        print("# Normal")
                        print("ok")
                    else:
                        print("# Sheet1")
                        print("42")
                    """
                ),
                encoding="utf-8",
            )
            os.chmod(fake_extract, 0o755)

            output = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT),
                    "--manifest",
                    str(manifest_path),
                    "--bin",
                    str(fake_extract),
                    "--show-skips",
                    "5",
                    "--min",
                    "0.0",
                ],
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn("files: 2", output.stdout)
            self.assertIn("rxls extracted: 2", output.stdout)
            self.assertIn("openpyxl-unreadable: 1", output.stdout)
            self.assertIn("comparable: 1", output.stdout)
            self.assertIn(
                "by_skip_decision: documented_oracle_limitation skipped=1",
                output.stdout,
            )
            self.assertIn(
                "by_skip_evidence: openpyxl_root_styles_path_variant skipped=1",
                output.stdout,
            )
            self.assertRegex(
                output.stdout,
                r"skip: kind=openpyxl-unreadable decision=documented_oracle_limitation evidence=openpyxl_root_styles_path_variant path=.*root-styles\.xlsx reason=UnsupportedOpenpyxlRootStylesPathVariant: openpyxl requires xl/styles\.xml but package uses styles\.xml",
            )

    def test_shared_string_amplification_is_documented_bounded_extraction(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            normal_path = base / "normal.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.title = "Normal"
            workbook.active["A1"] = "ok"
            workbook.save(normal_path)

            amplification_path = base / "shared-string-amplification.xlsx"
            _write_shared_string_amplification_workbook(
                amplification_path,
                shared_text="X" * 40,
                references=2,
            )

            manifest_path = base / "manifest.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "source": "unit",
                                "path": "normal.xlsx",
                                "local_path": str(normal_path),
                                "status": "downloaded",
                            },
                            {
                                "source": "unit",
                                "path": "shared-string-amplification.xlsx",
                                "local_path": str(amplification_path),
                                "status": "downloaded",
                            },
                        ]
                    }
                ),
                encoding="utf-8",
            )

            fake_extract = base / "fake-rxls-extract"
            fake_extract.write_text(
                textwrap.dedent(
                    """\
                    #!/usr/bin/env python3
                    import pathlib
                    if pathlib.Path(__import__("sys").argv[1]).name == "normal.xlsx":
                        print("# Normal")
                        print("ok")
                    else:
                        print("# Sheet1")
                        print("X" * 40 + "\\t" + "X" * 40)
                    """
                ),
                encoding="utf-8",
            )
            os.chmod(fake_extract, 0o755)

            output = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT),
                    "--manifest",
                    str(manifest_path),
                    "--bin",
                    str(fake_extract),
                    "--max-hash-chars",
                    "50",
                    "--show-skips",
                    "5",
                    "--min",
                    "0.0",
                ],
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn("files: 2", output.stdout)
            self.assertIn("rxls extracted: 2", output.stdout)
            self.assertIn("comparable: 1", output.stdout)
            self.assertIn("bounded-shared-string-expansions: 1", output.stdout)
            self.assertIn(
                "by_skip_decision: documented_bounded_extraction skipped=1",
                output.stdout,
            )
            self.assertIn(
                "by_skip_evidence: shared_string_expansion_budget_exceeded skipped=1",
                output.stdout,
            )
            self.assertRegex(
                output.stdout,
                r"skip: kind=bounded-shared-string-expansion decision=documented_bounded_extraction evidence=shared_string_expansion_budget_exceeded path=.*shared-string-amplification\.xlsx reason=expanded shared-string text 80 exceeds 50 hash budget",
            )

    def test_manifest_mode_selects_downloaded_xlsx_entries(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            workbook_path = base / "nested" / "oracle.xlsx"
            workbook_path.parent.mkdir()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Data"
            sheet["A1"] = "hello"
            sheet["B1"] = 2
            workbook.save(workbook_path)

            ignored_xls = base / "ignored.xls"
            ignored_xls.write_bytes(b"not part of xlsx oracle")
            missing_xlsx = base / "missing.xlsx"

            manifest_path = base / "manifest.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "source": "unit",
                                "path": "nested/oracle.xlsx",
                                "local_path": str(workbook_path),
                                "status": "downloaded",
                            },
                            {
                                "source": "unit",
                                "path": "ignored.xls",
                                "local_path": str(ignored_xls),
                                "status": "downloaded",
                            },
                            {
                                "source": "unit",
                                "path": "missing.xlsx",
                                "local_path": str(missing_xlsx),
                                "status": "failed",
                            },
                        ]
                    }
                ),
                encoding="utf-8",
            )

            fake_extract = base / "fake-rxls-extract"
            fake_extract.write_text(
                textwrap.dedent(
                    """\
                    #!/usr/bin/env python3
                    print("# Data")
                    print("hello\\t2")
                    """
                ),
                encoding="utf-8",
            )
            os.chmod(fake_extract, 0o755)

            output = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT),
                    "--manifest",
                    str(manifest_path),
                    "--bin",
                    str(fake_extract),
                    "--limit",
                    "10",
                    "--min",
                    "1.0",
                ],
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn(f"manifest: {manifest_path}", output.stdout)
            self.assertIn("files: 1", output.stdout)
            self.assertIn("rxls extracted: 1", output.stdout)
            self.assertIn("openpyxl-unreadable: 0", output.stdout)
            self.assertIn("rxls vs openpyxl: mean parity 100.000%", output.stdout)

    def test_oversized_exact_manifest_comparisons_use_hash_without_diffing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            small_path = base / "small.xlsx"
            small_workbook = openpyxl.Workbook()
            small_workbook.active.title = "Small"
            small_workbook.active["A1"] = "ok"
            small_workbook.save(small_path)

            large_path = base / "large.xlsx"
            large_workbook = openpyxl.Workbook()
            large_sheet = large_workbook.active
            large_sheet.title = "Large"
            large_sheet["A1"] = "x" * 200
            large_workbook.save(large_path)

            manifest_path = base / "manifest.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "source": "unit",
                                "path": "small.xlsx",
                                "local_path": str(small_path),
                                "status": "downloaded",
                            },
                            {
                                "source": "unit",
                                "path": "large.xlsx",
                                "local_path": str(large_path),
                                "status": "downloaded",
                            },
                        ]
                    }
                ),
                encoding="utf-8",
            )

            fake_extract = base / "fake-rxls-extract"
            fake_extract.write_text(
                textwrap.dedent(
                    """\
                    #!/usr/bin/env python3
                    import pathlib
                    if pathlib.Path(__import__("sys").argv[1]).name == "small.xlsx":
                        print("# Small")
                        print("ok")
                    else:
                        print("# Large")
                        print("x" * 200)
                    """
                ),
                encoding="utf-8",
            )
            os.chmod(fake_extract, 0o755)

            output = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT),
                    "--manifest",
                    str(manifest_path),
                    "--bin",
                    str(fake_extract),
                    "--max-compare-chars",
                    "50",
                    "--min",
                    "1.0",
                ],
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn("files: 2", output.stdout)
            self.assertIn("comparable: 2", output.stdout)
            self.assertIn("oversized-comparisons: 0", output.stdout)
            self.assertIn("hash-exact-comparisons: 1", output.stdout)
            self.assertIn("rxls vs openpyxl: mean parity 100.000%", output.stdout)

    def test_oversized_mismatch_manifest_comparisons_are_reported_and_skipped(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            small_path = base / "small.xlsx"
            small_workbook = openpyxl.Workbook()
            small_workbook.active.title = "Small"
            small_workbook.active["A1"] = "ok"
            small_workbook.save(small_path)

            large_path = base / "large.xlsx"
            large_workbook = openpyxl.Workbook()
            large_sheet = large_workbook.active
            large_sheet.title = "Large"
            large_sheet["A1"] = "x" * 200
            large_workbook.save(large_path)

            manifest_path = base / "manifest.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "source": "unit",
                                "path": "small.xlsx",
                                "local_path": str(small_path),
                                "status": "downloaded",
                            },
                            {
                                "source": "unit",
                                "path": "large.xlsx",
                                "local_path": str(large_path),
                                "status": "downloaded",
                            },
                        ]
                    }
                ),
                encoding="utf-8",
            )

            fake_extract = base / "fake-rxls-extract"
            fake_extract.write_text(
                textwrap.dedent(
                    """\
                    #!/usr/bin/env python3
                    import pathlib
                    if pathlib.Path(__import__("sys").argv[1]).name == "small.xlsx":
                        print("# Small")
                        print("ok")
                    else:
                        print("# Large")
                        print("y" * 200)
                    """
                ),
                encoding="utf-8",
            )
            os.chmod(fake_extract, 0o755)

            output = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT),
                    "--manifest",
                    str(manifest_path),
                    "--bin",
                    str(fake_extract),
                    "--max-compare-chars",
                    "50",
                    "--show-skips",
                    "10",
                    "--min",
                    "1.0",
                ],
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn("files: 2", output.stdout)
            self.assertIn("comparable: 1", output.stdout)
            self.assertIn("oversized-comparisons: 1", output.stdout)
            self.assertIn("hash-exact-comparisons: 0", output.stdout)
            self.assertRegex(
                output.stdout,
                r"skip: kind=oversized-comparison decision=needs_bounded_oracle evidence=comparison_budget_exceeded path=.*large\.xlsx reason=combined text length ",
            )

    def test_oversized_exact_hash_comparisons_respect_hash_budget(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            small_path = base / "small.xlsx"
            small_workbook = openpyxl.Workbook()
            small_workbook.active.title = "Small"
            small_workbook.active["A1"] = "ok"
            small_workbook.save(small_path)

            large_path = base / "large.xlsx"
            large_workbook = openpyxl.Workbook()
            large_sheet = large_workbook.active
            large_sheet.title = "Large"
            large_sheet["A1"] = "x" * 200
            large_workbook.save(large_path)

            manifest_path = base / "manifest.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "source": "unit",
                                "path": "small.xlsx",
                                "local_path": str(small_path),
                                "status": "downloaded",
                            },
                            {
                                "source": "unit",
                                "path": "large.xlsx",
                                "local_path": str(large_path),
                                "status": "downloaded",
                            },
                        ]
                    }
                ),
                encoding="utf-8",
            )

            fake_extract = base / "fake-rxls-extract"
            fake_extract.write_text(
                textwrap.dedent(
                    """\
                    #!/usr/bin/env python3
                    import pathlib
                    if pathlib.Path(__import__("sys").argv[1]).name == "small.xlsx":
                        print("# Small")
                        print("ok")
                    else:
                        print("# Large")
                        print("x" * 200)
                    """
                ),
                encoding="utf-8",
            )
            os.chmod(fake_extract, 0o755)

            output = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT),
                    "--manifest",
                    str(manifest_path),
                    "--bin",
                    str(fake_extract),
                    "--max-compare-chars",
                    "50",
                    "--max-hash-chars",
                    "50",
                    "--show-skips",
                    "10",
                    "--min",
                    "1.0",
                ],
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn("files: 2", output.stdout)
            self.assertIn("comparable: 1", output.stdout)
            self.assertIn("oversized-comparisons: 1", output.stdout)
            self.assertIn("hash-exact-comparisons: 0", output.stdout)
            self.assertRegex(
                output.stdout,
                r"skip: kind=oversized-comparison decision=needs_bounded_oracle evidence=comparison_budget_exceeded path=.*large\.xlsx reason=combined text length .* exceeds 50 hash budget",
            )

    def test_show_skips_reports_representative_skip_reasons(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            small_path = base / "small.xlsx"
            small_workbook = openpyxl.Workbook()
            small_workbook.active.title = "Small"
            small_workbook.active["A1"] = "ok"
            small_workbook.save(small_path)

            invalid_path = base / "invalid.xlsx"
            invalid_path.write_bytes(b"not an ooxml workbook")

            large_path = base / "large.xlsx"
            large_workbook = openpyxl.Workbook()
            large_sheet = large_workbook.active
            large_sheet.title = "Large"
            large_sheet["A1"] = "x" * 200
            large_workbook.save(large_path)

            sparse_path = base / "sparse.xlsx"
            sparse_workbook = openpyxl.Workbook()
            sparse_sheet = sparse_workbook.active
            sparse_sheet.title = "Sparse"
            sparse_sheet.cell(row=1000, column=1000, value="far")
            sparse_workbook.save(sparse_path)

            manifest_path = base / "manifest.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "source": "unit",
                                "path": "small.xlsx",
                                "local_path": str(small_path),
                                "status": "downloaded",
                            },
                            {
                                "source": "unit",
                                "path": "invalid.xlsx",
                                "local_path": str(invalid_path),
                                "status": "downloaded",
                            },
                            {
                                "source": "unit",
                                "path": "large.xlsx",
                                "local_path": str(large_path),
                                "status": "downloaded",
                            },
                            {
                                "source": "unit",
                                "path": "sparse.xlsx",
                                "local_path": str(sparse_path),
                                "status": "downloaded",
                            },
                        ]
                    }
                ),
                encoding="utf-8",
            )

            fake_extract = base / "fake-rxls-extract"
            fake_extract.write_text(
                textwrap.dedent(
                    """\
                    #!/usr/bin/env python3
                    import pathlib
                    name = pathlib.Path(__import__("sys").argv[1]).name
                    if name == "small.xlsx":
                        print("# Small")
                        print("ok")
                    elif name == "large.xlsx":
                        print("# Large")
                        print("y" * 200)
                    elif name == "sparse.xlsx":
                        print("# Sparse")
                        print("far")
                    else:
                        print("# Invalid")
                        print("rxls recovered text")
                    """
                ),
                encoding="utf-8",
            )
            os.chmod(fake_extract, 0o755)

            output = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT),
                    "--manifest",
                    str(manifest_path),
                    "--bin",
                    str(fake_extract),
                    "--max-compare-chars",
                    "50",
                    "--max-worksheet-cells",
                    "100",
                    "--show-skips",
                    "10",
                    "--min",
                    "1.0",
                ],
                capture_output=True,
                text=True,
            )

            self.assertEqual(output.returncode, 0, output.stdout + output.stderr)
            self.assertIn("files: 4", output.stdout)
            self.assertIn("comparable: 2", output.stdout)
            self.assertIn("openpyxl-unreadable: 1", output.stdout)
            self.assertIn("oversized-comparisons: 1", output.stdout)
            self.assertIn("oversized-worksheets: 0", output.stdout)
            self.assertIn(
                "by_skip_decision: needs_corpus_crosscheck skipped=1",
                output.stdout,
            )
            self.assertIn(
                "by_skip_decision: needs_bounded_oracle skipped=1",
                output.stdout,
            )
            self.assertIn(
                "by_skip_evidence: openpyxl_non_zip_container skipped=1",
                output.stdout,
            )
            self.assertIn(
                "by_skip_evidence: comparison_budget_exceeded skipped=1",
                output.stdout,
            )
            self.assertRegex(
                output.stdout,
                r"skip: kind=openpyxl-unreadable decision=needs_corpus_crosscheck evidence=openpyxl_non_zip_container path=.*invalid\.xlsx reason=BadZipFile:",
            )
            self.assertRegex(
                output.stdout,
                r"skip: kind=oversized-comparison decision=needs_bounded_oracle evidence=comparison_budget_exceeded path=.*large\.xlsx reason=combined text length ",
            )

    def test_show_skips_refines_non_zip_rows_from_corpus_report(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            small_path = base / "small.xlsx"
            small_workbook = openpyxl.Workbook()
            small_workbook.active.title = "Small"
            small_workbook.active["A1"] = "ok"
            small_workbook.save(small_path)

            protected_path = base / "apache-poi" / "test-data" / "spreadsheet" / "protected.xlsx"
            protected_path.parent.mkdir(parents=True)
            protected_path.write_bytes(b"not a zip container")

            manifest_path = base / "manifest.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "source": "unit",
                                "path": "small.xlsx",
                                "local_path": str(small_path),
                                "status": "downloaded",
                            },
                            {
                                "source": "unit",
                                "path": "test-data/spreadsheet/protected.xlsx",
                                "local_path": str(protected_path),
                                "status": "downloaded",
                            },
                        ]
                    }
                ),
                encoding="utf-8",
            )

            corpus_report = base / "corpus-report.txt"
            corpus_report.write_text(
                "failure: .xlsx test-data/spreadsheet/protected.xlsx "
                "kind=unsupported_encrypted_ooxml decision=unsupported_encrypted "
                "evidence=encrypted_ooxml_package container=ole2 "
                "extension_mismatch=true parse: unsupported encrypted OOXML package\n",
                encoding="utf-8",
            )

            fake_extract = base / "fake-rxls-extract"
            fake_extract.write_text(
                textwrap.dedent(
                    """\
                    #!/usr/bin/env python3
                    import pathlib
                    name = pathlib.Path(__import__("sys").argv[1]).name
                    if name == "small.xlsx":
                        print("# Small")
                        print("ok")
                    """
                ),
                encoding="utf-8",
            )
            os.chmod(fake_extract, 0o755)

            output = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT),
                    "--manifest",
                    str(manifest_path),
                    "--bin",
                    str(fake_extract),
                    "--corpus-report",
                    str(corpus_report),
                    "--show-skips",
                    "10",
                    "--min",
                    "1.0",
                ],
                capture_output=True,
                text=True,
            )

            self.assertEqual(output.returncode, 0, output.stdout + output.stderr)
            self.assertIn("files: 2", output.stdout)
            self.assertIn("comparable: 1", output.stdout)
            self.assertIn("openpyxl-unreadable: 1", output.stdout)
            self.assertIn(
                "by_skip_decision: unsupported_encrypted skipped=1",
                output.stdout,
            )
            self.assertIn(
                "by_skip_evidence: encrypted_ooxml_package skipped=1",
                output.stdout,
            )
            self.assertIn(
                "by_skip_corpus_kind: unsupported_encrypted_ooxml skipped=1",
                output.stdout,
            )
            self.assertNotIn("needs_corpus_crosscheck", output.stdout)
            self.assertRegex(
                output.stdout,
                r"skip: kind=openpyxl-unreadable decision=unsupported_encrypted evidence=encrypted_ooxml_package corpus_kind=unsupported_encrypted_ooxml path=.*protected\.xlsx reason=BadZipFile:",
            )

    def test_show_worst_reports_lowest_parity_manifest_entries(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            perfect_path = base / "perfect.xlsx"
            perfect_workbook = openpyxl.Workbook()
            perfect_workbook.active.title = "Perfect"
            perfect_workbook.active["A1"] = "match"
            perfect_workbook.save(perfect_path)

            mismatch_path = base / "mismatch.xlsx"
            mismatch_workbook = openpyxl.Workbook()
            mismatch_workbook.active.title = "Mismatch"
            mismatch_workbook.active["A1"] = "oracle"
            mismatch_workbook.save(mismatch_path)

            manifest_path = base / "manifest.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "source": "unit",
                                "path": "perfect.xlsx",
                                "local_path": str(perfect_path),
                                "status": "downloaded",
                            },
                            {
                                "source": "unit",
                                "path": "mismatch.xlsx",
                                "local_path": str(mismatch_path),
                                "status": "downloaded",
                            },
                        ]
                    }
                ),
                encoding="utf-8",
            )

            fake_extract = base / "fake-rxls-extract"
            fake_extract.write_text(
                textwrap.dedent(
                    """\
                    #!/usr/bin/env python3
                    import pathlib
                    if pathlib.Path(__import__("sys").argv[1]).name == "perfect.xlsx":
                        print("# Perfect")
                        print("match")
                    else:
                        print("# Mismatch")
                        print("rxls")
                    """
                ),
                encoding="utf-8",
            )
            os.chmod(fake_extract, 0o755)

            output = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT),
                    "--manifest",
                    str(manifest_path),
                    "--bin",
                    str(fake_extract),
                    "--show-worst",
                    "1",
                    "--min",
                    "0.0",
                ],
                capture_output=True,
                text=True,
            )

            self.assertEqual(output.returncode, 0, output.stderr)
            self.assertIn("low-parity:", output.stdout)
            low_parity_lines = [
                line
                for line in output.stdout.splitlines()
                if line.startswith("low-parity:")
            ]
            self.assertEqual(len(low_parity_lines), 1)
            self.assertIn("mismatch.xlsx", low_parity_lines[0])
            self.assertNotIn("perfect.xlsx", low_parity_lines[0])

    def test_manifest_uses_bounded_sparse_cells_without_iterating_declared_grid(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            small_path = base / "small.xlsx"
            small_workbook = openpyxl.Workbook()
            small_workbook.active.title = "Small"
            small_workbook.active["A1"] = "ok"
            small_workbook.save(small_path)

            sparse_path = base / "sparse.xlsx"
            sparse_workbook = openpyxl.Workbook()
            sparse_sheet = sparse_workbook.active
            sparse_sheet.title = "Sparse"
            sparse_sheet.cell(row=1000, column=1000, value="far")
            sparse_workbook.save(sparse_path)

            manifest_path = base / "manifest.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "source": "unit",
                                "path": "small.xlsx",
                                "local_path": str(small_path),
                                "status": "downloaded",
                            },
                            {
                                "source": "unit",
                                "path": "sparse.xlsx",
                                "local_path": str(sparse_path),
                                "status": "downloaded",
                            },
                        ]
                    }
                ),
                encoding="utf-8",
            )

            fake_extract = base / "fake-rxls-extract"
            fake_extract.write_text(
                textwrap.dedent(
                    """\
                    #!/usr/bin/env python3
                    import pathlib
                    if pathlib.Path(__import__("sys").argv[1]).name == "small.xlsx":
                        print("# Small")
                        print("ok")
                    else:
                        print("# Sparse")
                        print("far")
                    """
                ),
                encoding="utf-8",
            )
            os.chmod(fake_extract, 0o755)

            output = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT),
                    "--manifest",
                    str(manifest_path),
                    "--bin",
                    str(fake_extract),
                    "--max-worksheet-cells",
                    "100",
                    "--min",
                    "1.0",
                ],
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn("files: 2", output.stdout)
            self.assertIn("comparable: 2", output.stdout)
            self.assertIn("oversized-worksheets: 0", output.stdout)
            self.assertIn("rxls vs openpyxl: mean parity 100.000%", output.stdout)

    def test_manifest_skips_legacy_beta_namespace_workbooks(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            normal_path = base / "normal.xlsx"
            normal_workbook = openpyxl.Workbook()
            normal_workbook.active.title = "Normal"
            normal_workbook.active["A1"] = "ok"
            normal_workbook.save(normal_path)

            beta_path = base / "beta.xlsx"
            with ZipFile(beta_path, "w", ZIP_DEFLATED) as package:
                package.writestr(
                    "[Content_Types].xml",
                    textwrap.dedent(
                        """\
                        <?xml version="1.0" encoding="UTF-8"?>
                        <Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
                          <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
                          <Default Extension="xml" ContentType="application/xml"/>
                          <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
                          <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
                        </Types>
                        """
                    ),
                )
                package.writestr(
                    "_rels/.rels",
                    textwrap.dedent(
                        """\
                        <?xml version="1.0" encoding="UTF-8"?>
                        <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
                          <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
                        </Relationships>
                        """
                    ),
                )
                package.writestr(
                    "xl/_rels/workbook.xml.rels",
                    textwrap.dedent(
                        """\
                        <?xml version="1.0" encoding="UTF-8"?>
                        <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
                          <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
                        </Relationships>
                        """
                    ),
                )
                package.writestr(
                    "xl/workbook.xml",
                    textwrap.dedent(
                        """\
                        <?xml version="1.0" encoding="UTF-8"?>
                        <workbook xmlns="http://schemas.microsoft.com/office/excel/2006/2" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
                          <sheets>
                            <sheet name="Beta" sheetId="1" r:id="rId1"/>
                          </sheets>
                        </workbook>
                        """
                    ),
                )
                package.writestr(
                    "xl/worksheets/sheet1.xml",
                    textwrap.dedent(
                        """\
                        <?xml version="1.0" encoding="UTF-8"?>
                        <worksheet xmlns="http://schemas.microsoft.com/office/excel/2006/2" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
                          <dimension ref="A1"/>
                          <sheetData>
                            <row r="1"><c r="A1"><v>42</v></c></row>
                          </sheetData>
                        </worksheet>
                        """
                    ),
                )

            manifest_path = base / "manifest.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "source": "unit",
                                "path": "normal.xlsx",
                                "local_path": str(normal_path),
                                "status": "downloaded",
                            },
                            {
                                "source": "unit",
                                "path": "beta.xlsx",
                                "local_path": str(beta_path),
                                "status": "downloaded",
                            },
                        ]
                    }
                ),
                encoding="utf-8",
            )

            fake_extract = base / "fake-rxls-extract"
            fake_extract.write_text(
                textwrap.dedent(
                    """\
                    #!/usr/bin/env python3
                    import pathlib
                    if pathlib.Path(__import__("sys").argv[1]).name == "normal.xlsx":
                        print("# Normal")
                        print("ok")
                    else:
                        print("# Beta")
                        print("42")
                    """
                ),
                encoding="utf-8",
            )
            os.chmod(fake_extract, 0o755)

            output = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT),
                    "--manifest",
                    str(manifest_path),
                    "--bin",
                    str(fake_extract),
                    "--show-worst",
                    "2",
                    "--min",
                    "1.0",
                ],
                check=True,
                capture_output=True,
                text=True,
            )

            self.assertIn("files: 2", output.stdout)
            self.assertIn("openpyxl-unreadable: 1", output.stdout)
            self.assertIn("comparable: 1", output.stdout)
            self.assertIn("rxls vs openpyxl: mean parity 100.000%", output.stdout)
            self.assertNotIn("beta.xlsx", output.stdout)

    def test_manifest_skips_malformed_inline_string_value_workbooks(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            normal_path = base / "normal.xlsx"
            normal_workbook = openpyxl.Workbook()
            normal_workbook.active.title = "Normal"
            normal_workbook.active["A1"] = "ok"
            normal_workbook.save(normal_path)

            malformed_path = base / "malformed.xlsx"
            with ZipFile(malformed_path, "w", ZIP_DEFLATED) as package:
                package.writestr(
                    "[Content_Types].xml",
                    textwrap.dedent(
                        """\
                        <?xml version="1.0" encoding="UTF-8"?>
                        <Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
                          <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
                          <Default Extension="xml" ContentType="application/xml"/>
                          <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
                          <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
                        </Types>
                        """
                    ),
                )
                package.writestr(
                    "_rels/.rels",
                    textwrap.dedent(
                        """\
                        <?xml version="1.0" encoding="UTF-8"?>
                        <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
                          <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
                        </Relationships>
                        """
                    ),
                )
                package.writestr(
                    "xl/_rels/workbook.xml.rels",
                    textwrap.dedent(
                        """\
                        <?xml version="1.0" encoding="UTF-8"?>
                        <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
                          <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
                        </Relationships>
                        """
                    ),
                )
                package.writestr(
                    "xl/workbook.xml",
                    textwrap.dedent(
                        """\
                        <?xml version="1.0" encoding="UTF-8"?>
                        <workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
                          <sheets>
                            <sheet name="Malformed" sheetId="1" r:id="rId1"/>
                          </sheets>
                        </workbook>
                        """
                    ),
                )
                package.writestr(
                    "xl/worksheets/sheet1.xml",
                    textwrap.dedent(
                        """\
                        <?xml version="1.0" encoding="UTF-8"?>
                        <worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
                          <dimension ref="A1"/>
                          <sheetData>
                            <row r="1">
                              <c r="A1" t="inlineStr"><v>lost by openpyxl</v></c>
                            </row>
                          </sheetData>
                        </worksheet>
                        """
                    ),
                )

            manifest_path = base / "manifest.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "source": "unit",
                                "path": "normal.xlsx",
                                "local_path": str(normal_path),
                                "status": "downloaded",
                            },
                            {
                                "source": "unit",
                                "path": "malformed.xlsx",
                                "local_path": str(malformed_path),
                                "status": "downloaded",
                            },
                        ]
                    }
                ),
                encoding="utf-8",
            )

            fake_extract = base / "fake-rxls-extract"
            fake_extract.write_text(
                textwrap.dedent(
                    """\
                    #!/usr/bin/env python3
                    import pathlib
                    if pathlib.Path(__import__("sys").argv[1]).name == "normal.xlsx":
                        print("# Normal")
                        print("ok")
                    else:
                        print("# Malformed")
                        print("lost by openpyxl")
                    """
                ),
                encoding="utf-8",
            )
            os.chmod(fake_extract, 0o755)

            output = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT),
                    "--manifest",
                    str(manifest_path),
                    "--bin",
                    str(fake_extract),
                    "--show-worst",
                    "2",
                    "--min",
                    "1.0",
                ],
                capture_output=True,
                text=True,
            )

            self.assertEqual(output.returncode, 0, output.stdout + output.stderr)
            self.assertIn("files: 2", output.stdout)
            self.assertIn("openpyxl-unreadable: 1", output.stdout)
            self.assertIn("comparable: 1", output.stdout)
            self.assertIn("rxls vs openpyxl: mean parity 100.000%", output.stdout)
            self.assertNotIn("malformed.xlsx", output.stdout)

    def test_xlsm_oracle_ignores_macro_sheets_for_worksheet_text_parity(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            workbook_path = base / "macro.xlsm"
            with ZipFile(workbook_path, "w", ZIP_DEFLATED) as package:
                package.writestr(
                    "[Content_Types].xml",
                    textwrap.dedent(
                        """\
                        <?xml version="1.0" encoding="UTF-8"?>
                        <Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
                          <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
                          <Default Extension="xml" ContentType="application/xml"/>
                          <Override PartName="/xl/workbook.xml" ContentType="application/vnd.ms-excel.sheet.macroEnabled.main+xml"/>
                          <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
                          <Override PartName="/xl/macrosheets/sheet1.xml" ContentType="application/vnd.ms-excel.macrosheet+xml"/>
                        </Types>
                        """
                    ),
                )
                package.writestr(
                    "_rels/.rels",
                    textwrap.dedent(
                        """\
                        <?xml version="1.0" encoding="UTF-8"?>
                        <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
                          <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
                        </Relationships>
                        """
                    ),
                )
                package.writestr(
                    "xl/workbook.xml",
                    textwrap.dedent(
                        """\
                        <?xml version="1.0" encoding="UTF-8"?>
                        <workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
                          <sheets>
                            <sheet name="Data" sheetId="1" r:id="rId1"/>
                            <sheet name="Macro Sheet" sheetId="2" r:id="rId2"/>
                          </sheets>
                        </workbook>
                        """
                    ),
                )
                package.writestr(
                    "xl/_rels/workbook.xml.rels",
                    textwrap.dedent(
                        """\
                        <?xml version="1.0" encoding="UTF-8"?>
                        <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
                          <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
                          <Relationship Id="rId2" Type="http://schemas.microsoft.com/office/2006/relationships/xlMacrosheet" Target="macrosheets/sheet1.xml"/>
                        </Relationships>
                        """
                    ),
                )
                package.writestr(
                    "xl/worksheets/sheet1.xml",
                    textwrap.dedent(
                        """\
                        <?xml version="1.0" encoding="UTF-8"?>
                        <worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
                          <sheetData>
                            <row r="1"><c r="A1" t="inlineStr"><is><t>ok</t></is></c></row>
                          </sheetData>
                        </worksheet>
                        """
                    ),
                )
                package.writestr(
                    "xl/macrosheets/sheet1.xml",
                    textwrap.dedent(
                        """\
                        <?xml version="1.0" encoding="UTF-8"?>
                        <xm:macrosheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:xm="http://schemas.microsoft.com/office/excel/2006/main">
                          <sheetData>
                            <row r="1"><c r="A1" t="inlineStr"><is><t>macro text</t></is></c></row>
                          </sheetData>
                        </xm:macrosheet>
                        """
                    ),
                )

            manifest_path = base / "manifest.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "source": "unit",
                                "path": "macro.xlsm",
                                "local_path": str(workbook_path),
                                "status": "downloaded",
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )

            fake_extract = base / "fake-rxls-extract"
            fake_extract.write_text(
                textwrap.dedent(
                    """\
                    #!/usr/bin/env python3
                    print("# Data")
                    print("ok")
                    """
                ),
                encoding="utf-8",
            )
            os.chmod(fake_extract, 0o755)

            output = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT),
                    "--manifest",
                    str(manifest_path),
                    "--bin",
                    str(fake_extract),
                    "--show-worst",
                    "1",
                    "--min",
                    "1.0",
                ],
                capture_output=True,
                text=True,
            )

            self.assertEqual(output.returncode, 0, output.stdout + output.stderr)
            self.assertIn("files: 1", output.stdout)
            self.assertIn("comparable: 1", output.stdout)
            self.assertIn("rxls vs openpyxl: mean parity 100.000%", output.stdout)

    def test_timedelta_oracle_uses_elapsed_hour_format(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            workbook_path = base / "duration.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Durations"
            sheet["A1"] = datetime.timedelta(days=10, hours=15, minutes=10, seconds=10)
            sheet["A1"].number_format = "[hh]:mm:ss"
            workbook.save(workbook_path)

            manifest_path = base / "manifest.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "source": "unit",
                                "path": "duration.xlsx",
                                "local_path": str(workbook_path),
                                "status": "downloaded",
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )

            fake_extract = base / "fake-rxls-extract"
            fake_extract.write_text(
                textwrap.dedent(
                    """\
                    #!/usr/bin/env python3
                    print("# Durations")
                    print("255:10:10")
                    """
                ),
                encoding="utf-8",
            )
            os.chmod(fake_extract, 0o755)

            output = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT),
                    "--manifest",
                    str(manifest_path),
                    "--bin",
                    str(fake_extract),
                    "--min",
                    "1.0",
                ],
                capture_output=True,
                text=True,
            )

            self.assertEqual(output.returncode, 0, output.stdout + output.stderr)
            self.assertIn("files: 1", output.stdout)
            self.assertIn("rxls vs openpyxl: mean parity 100.000%", output.stdout)

    def test_numeric_oracle_uses_conditional_elapsed_hour_format(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            workbook_path = base / "numeric-duration.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Durations"
            sheet["A1"] = 0.020833333333333332
            sheet["A1"].number_format = "[=0]?;[<4.16666666666667][hh]:mm:ss;[hh]:mm"
            workbook.save(workbook_path)

            manifest_path = base / "manifest.json"
            manifest_path.write_text(
                json.dumps(
                    {
                        "files": [
                            {
                                "source": "unit",
                                "path": "numeric-duration.xlsx",
                                "local_path": str(workbook_path),
                                "status": "downloaded",
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )

            fake_extract = base / "fake-rxls-extract"
            fake_extract.write_text(
                textwrap.dedent(
                    """\
                    #!/usr/bin/env python3
                    print("# Durations")
                    print("0:30:00")
                    """
                ),
                encoding="utf-8",
            )
            os.chmod(fake_extract, 0o755)

            output = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT),
                    "--manifest",
                    str(manifest_path),
                    "--bin",
                    str(fake_extract),
                    "--min",
                    "1.0",
                ],
                capture_output=True,
                text=True,
            )

            self.assertEqual(output.returncode, 0, output.stdout + output.stderr)
            self.assertIn("files: 1", output.stdout)
            self.assertIn("rxls vs openpyxl: mean parity 100.000%", output.stdout)


def _write_shared_strings_path_variant_workbook(
    path: Path,
    shared_strings_part: str,
    *,
    include_content_types: bool = True,
) -> None:
    with ZipFile(path, "w", ZIP_DEFLATED) as package:
        if include_content_types:
            package.writestr(
                "[Content_Types].xml",
                """<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>
</Types>
""",
            )
        package.writestr(
            "_rels/.rels",
            """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>
""",
        )
        package.writestr(
            "xl/workbook.xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
          xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="Sheet1" sheetId="1" r:id="rId1"/>
  </sheets>
</workbook>
""",
        )
        package.writestr(
            "xl/_rels/workbook.xml.rels",
            """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>
</Relationships>
""",
        )
        package.writestr(
            "xl/worksheets/sheet1.xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetData>
    <row r="1">
      <c r="A1" t="s"><v>0</v></c>
    </row>
  </sheetData>
</worksheet>
""",
        )
        package.writestr(
            shared_strings_part,
            """<?xml version="1.0" encoding="UTF-8"?>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="1" uniqueCount="1">
  <si><t>hello</t></si>
</sst>
""",
        )


def _write_empty_style_index_workbook(path: Path) -> None:
    with ZipFile(path, "w", ZIP_DEFLATED) as package:
        package.writestr(
            "[Content_Types].xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>
""",
        )
        package.writestr(
            "_rels/.rels",
            """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>
""",
        )
        package.writestr(
            "xl/workbook.xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
          xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="Sheet1" sheetId="1" r:id="rId1"/>
  </sheets>
</workbook>
""",
        )
        package.writestr(
            "xl/_rels/workbook.xml.rels",
            """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>
""",
        )
        package.writestr(
            "xl/styles.xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>
  <fills count="1"><fill><patternFill patternType="none"/></fill></fills>
  <borders count="1"><border/></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>
</styleSheet>
""",
        )
        package.writestr(
            "xl/worksheets/sheet1.xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetData>
    <row r="1">
      <c r="A1" s="" t="inlineStr"><is><t>empty</t></is></c>
    </row>
  </sheetData>
</worksheet>
""",
        )


def _write_root_styles_path_variant_workbook(path: Path) -> None:
    with ZipFile(path, "w", ZIP_DEFLATED) as package:
        package.writestr(
            "[Content_Types].xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>
""",
        )
        package.writestr(
            "_rels/.rels",
            """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="workbook.xml"/>
</Relationships>
""",
        )
        package.writestr(
            "workbook.xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
          xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="Sheet1" sheetId="1" r:id="rId1"/>
  </sheets>
</workbook>
""",
        )
        package.writestr(
            "_rels/workbook.xml.rels",
            """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>
""",
        )
        package.writestr(
            "styles.xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>
  <fills count="1"><fill><patternFill patternType="none"/></fill></fills>
  <borders count="1"><border/></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="2">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
  </cellXfs>
</styleSheet>
""",
        )
        package.writestr(
            "sheet1.xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetData>
    <row r="1">
      <c r="A1" s="1"><v>42</v></c>
    </row>
  </sheetData>
</worksheet>
""",
        )


def _write_shared_string_amplification_workbook(
    path: Path,
    *,
    shared_text: str,
    references: int,
) -> None:
    cells = "\n".join(
        f'      <c r="{chr(ord("A") + idx)}1" t="s"><v>0</v></c>'
        for idx in range(references)
    )
    with ZipFile(path, "w", ZIP_DEFLATED) as package:
        package.writestr(
            "[Content_Types].xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>
</Types>
""",
        )
        package.writestr(
            "_rels/.rels",
            """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>
""",
        )
        package.writestr(
            "xl/workbook.xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
          xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="Sheet1" sheetId="1" r:id="rId1"/>
  </sheets>
</workbook>
""",
        )
        package.writestr(
            "xl/_rels/workbook.xml.rels",
            """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>
</Relationships>
""",
        )
        package.writestr(
            "xl/worksheets/sheet1.xml",
            f"""<?xml version="1.0" encoding="UTF-8"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetData>
    <row r="1">
{cells}
    </row>
  </sheetData>
</worksheet>
""",
        )
        package.writestr(
            "xl/sharedStrings.xml",
            f"""<?xml version="1.0" encoding="UTF-8"?>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="1" uniqueCount="1">
  <si><t>{shared_text}</t></si>
</sst>
""",
        )


def _load_xlsx_openpyxl_parity():
    scripts_dir = str(ROOT / "scripts")
    if scripts_dir not in sys.path:
        sys.path.insert(0, scripts_dir)
    spec = importlib.util.spec_from_file_location("xlsx_openpyxl_parity", SCRIPT)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


if __name__ == "__main__":
    unittest.main()
