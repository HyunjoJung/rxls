#!/usr/bin/env python3
"""실전 public-corpus head-to-head: rxls vs calamine vs the oracle (xlrd/openpyxl).

For each file: run the oracle (ground-truth values), rxls, and calamine; record
coverage (extracted / open-fail / panic) and accuracy (multiset value recall vs the
oracle). Aggregate an honest scorecard.

    python3 compare-corpus.py <dir-of-files> <xls|xlsx> <rxls-bin> <cal-bin>
"""
import glob
import os
import subprocess
import sys
from collections import Counter

KIND = sys.argv[2]
RXLS = sys.argv[3]
CAL = sys.argv[4]
files = sorted(glob.glob(os.path.join(sys.argv[1], "*." + KIND)))


import datetime
import re

_EPOCH = datetime.date(1899, 12, 30)


def canon(t):
    # Date-fair: rxls renders dates as ISO (`2024-03-15`); xlrd/calamine emit the
    # raw Excel serial. Normalize an ISO date to its serial so the human-readable
    # rendering isn't scored as a miss.
    m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", t)
    if m:
        try:
            d = datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            return repr(float((d - _EPOCH).days))
        except ValueError:
            pass
    try:
        return repr(float(t))
    except ValueError:
        return t


def toks(text):
    out = []
    for line in text.splitlines():
        if line.startswith("#") or line.startswith("FAIL") or line == "PANIC":
            continue
        for t in line.split("\t"):
            t = t.strip()
            if t:
                out.append(canon(t.lower()))
    return out


def oracle(path):
    if KIND == "xls":
        import xlrd
        try:
            wb = xlrd.open_workbook(path, formatting_info=False)
        except Exception:
            return None
        out = []
        for sh in wb.sheets():
            for r in range(sh.nrows):
                for c in range(sh.ncols):
                    v = sh.cell_value(r, c)
                    if v == "" or v is None:
                        continue
                    if isinstance(v, float) and v.is_integer():
                        v = int(v)
                    out.append(canon(str(v).strip().lower()))
        return out
    else:
        from openpyxl import load_workbook
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return None
        out = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    if v is None or v == "":
                        continue
                    if isinstance(v, float) and v.is_integer():
                        v = int(v)
                    out.append(canon(str(v).strip().lower()))
        return out


def run(binp, path):
    try:
        p = subprocess.run([binp, path], capture_output=True, text=True,
                           encoding="utf-8", errors="replace", timeout=60)
        return p.returncode, p.stdout
    except subprocess.TimeoutExpired:
        return -9, ""


def recall(orc, got):
    if not orc:
        return None
    co, cg = Counter(orc), Counter(got)
    return sum(min(n, cg[k]) for k, n in co.items()) / len(orc)


n = 0
rxls_ext = cal_ext = cal_panic = rxls_fail = 0
rxls_only = cal_only = 0          # extracted where the OTHER tool failed
orc_unreadable_rxls = orc_unreadable_cal = 0  # oracle failed but tool extracted
rxls_recalls, cal_recalls = [], []
for f in files:
    n += 1
    orc = oracle(f)
    rxc, rxo = run(RXLS, f)
    cac, cao = run(CAL, f)
    rx_ok = rxc == 0 and not rxo.startswith("FAIL") and len(toks(rxo)) > 0
    ca_ok = cac == 0 and "PANIC" not in cao and not cao.lstrip().startswith("FAIL") and len(toks(cao)) > 0
    if rx_ok:
        rxls_ext += 1
    else:
        rxls_fail += 1
    if ca_ok:
        cal_ext += 1
    if cac == 101 or "PANIC" in cao:
        cal_panic += 1
    if rx_ok and not ca_ok:
        rxls_only += 1
    if ca_ok and not rx_ok:
        cal_only += 1
    if orc:
        if rx_ok:
            rr = recall(orc, toks(rxo))
            if rr is not None:
                rxls_recalls.append(rr)
        if ca_ok:
            cr = recall(orc, toks(cao))
            if cr is not None:
                cal_recalls.append(cr)
    else:
        if rx_ok:
            orc_unreadable_rxls += 1
        if ca_ok:
            orc_unreadable_cal += 1
    if n % 50 == 0:
        print(f"  ...{n}/{len(files)}", file=sys.stderr)


def mean(xs):
    return sum(xs) / len(xs) if xs else 0.0


print(f"\n=== {sys.argv[1]} ({KIND}) - {n} files ===")
print(f"rxls extracted:      {rxls_ext}/{n}   (fail/empty: {rxls_fail})")
print(f"calamine extracted:  {cal_ext}/{n}   (panics: {cal_panic})")
print(f"rxls-only (calamine failed):  {rxls_only}")
print(f"calamine-only (rxls failed):  {cal_only}")
print(f"oracle-unreadable but extracted: rxls {orc_unreadable_rxls}, calamine {orc_unreadable_cal}")
print(f"mean recall vs oracle: rxls {mean(rxls_recalls):.3%} ({len(rxls_recalls)})  |  calamine {mean(cal_recalls):.3%} ({len(cal_recalls)})")
