#!/usr/bin/env python3
"""Reproducible parity harness: rxls vs openpyxl for OOXML `.xlsx`/`.xlsm`.

openpyxl (the canonical Python OOXML reader) is used as an independent oracle —
NOT the rxls golden — so this measures rxls against a real reference parser. For
each `.xlsx`/`.xlsm` in the corpus it renders an openpyxl "golden" text the same
way rxls does (cells in row/col order; dates as ISO; literal-aware percent ×100;
numbers whole->int) and compares it to the rxls `extract` example output with a
whitespace-insensitive ratio. It mirrors `scripts/xls-xlrd-parity.py` closely,
swapping xlrd→openpyxl, so the two harnesses apply identical rendering rules.

Usage:
    python scripts/xlsx-openpyxl-parity.py \
        --corpus local/xls-poc/ooxml_pub \
        --bin    local/xls-poc/rxls-extract.exe

    cargo run --features full --bin rxls -- \
        corpus-report local/public-corpus/manifest.json --limit 200 \
        > local/public-corpus/corpus-report.txt
    python scripts/xlsx-openpyxl-parity.py \
        --manifest local/public-corpus/manifest.json \
        --bin      target/debug/examples/extract \
        --corpus-report local/public-corpus/corpus-report.txt \
        --limit    50 \
        --show-skips 20

Requires: pip install openpyxl . Exits non-zero if mean parity < --min (0.95).
The string comparison is bounded by `--max-compare-chars` so public-corpus
workbooks with very large extracted text do not make normalization or diffing
hang. Oversized but exact-normalized outputs can still be admitted through a
bounded SHA-256 check capped by `--max-hash-chars`. Openpyxl worksheet iteration
is also bounded by `--max-worksheet-cells`; sparse sheets with huge declared
dimensions are rendered from openpyxl's loaded cell map when that map is
bounded, and skipped only when the loaded map itself exceeds the budget.
Shared-string amplification is preflighted against `--max-hash-chars` so a small
package that clones a large shared string into many cells is documented as
bounded extraction without materializing the full oracle text.
`--show-skips` prints bounded file-level skip lines with oracle decision/evidence
tags.
"""
import argparse
import datetime
import difflib
import hashlib
import io
import posixpath
import re
import subprocess
import sys
import zipfile
from xml.etree import ElementTree as ET

from public_corpus_manifest import (
    corpus_files,
    emit_parity_provenance,
    manifest_files,
    report_path,
    report_reason,
    report_source_root,
    resolve_binary,
)


class OversizedWorksheet(Exception):
    """Raised when an openpyxl sheet declares too many cells to iterate safely."""


class UnsupportedOpenpyxlNamespace(Exception):
    """Raised when openpyxl silently drops workbook data for a known namespace."""


class UnsupportedOpenpyxlInlineStringValue(Exception):
    """Raised when openpyxl silently drops malformed inline-string values."""


class UnsupportedOpenpyxlSharedStringsPathVariant(Exception):
    """Raised when openpyxl cannot resolve a sharedStrings part path variant."""


class UnsupportedOpenpyxlEmptyStyleIndex(Exception):
    """Raised when openpyxl cannot handle empty cell style index attributes."""


class UnsupportedOpenpyxlRootStylesPathVariant(Exception):
    """Raised when openpyxl cannot resolve a non-canonical styles part path."""


LEGACY_BETA_SPREADSHEETML_NS = "http://schemas.microsoft.com/office/excel/2006/2"


def _local_name(tag):
    return tag.rsplit("}", 1)[-1] if tag.startswith("{") else tag


def _attr_local(elem, name):
    for key, value in elem.attrib.items():
        if _local_name(key) == name:
            return value
    return None


def _workbook_root_namespace(path):
    try:
        with zipfile.ZipFile(path) as package:
            workbook = package.read("xl/workbook.xml")
    except Exception:
        return None
    try:
        for _, elem in ET.iterparse(io.BytesIO(workbook), events=("start",)):
            tag = elem.tag
            return tag[1:].split("}", 1)[0] if tag.startswith("{") else ""
    except ET.ParseError:
        return None
    return None


def _worksheet_titles(path):
    try:
        with zipfile.ZipFile(path) as package:
            workbook = package.read("xl/workbook.xml")
            workbook_rels = package.read("xl/_rels/workbook.xml.rels")
    except Exception:
        return None
    try:
        rel_types = {}
        rel_root = ET.fromstring(workbook_rels)
        for rel in rel_root.iter():
            if _local_name(rel.tag) != "Relationship":
                continue
            rel_id = rel.attrib.get("Id")
            rel_type = rel.attrib.get("Type", "")
            if rel_id:
                rel_types[rel_id] = rel_type.rsplit("/", 1)[-1].lower()

        titles = set()
        workbook_root = ET.fromstring(workbook)
        for sheet in workbook_root.iter():
            if _local_name(sheet.tag) != "sheet":
                continue
            title = sheet.attrib.get("name")
            rid = _attr_local(sheet, "id")
            kind = rel_types.get(rid or "", "worksheet")
            if kind == "worksheet":
                titles.add(title)
        return titles
    except ET.ParseError:
        return None


def _worksheet_has_inline_string_v_value(stream):
    depth = 0
    inline_cell_depth = None
    inline_cell_has_is = False
    inline_cell_has_v = False
    for event, elem in ET.iterparse(stream, events=("start", "end")):
        if event == "start":
            depth += 1
            name = _local_name(elem.tag)
            if name == "c" and elem.attrib.get("t") == "inlineStr":
                inline_cell_depth = depth
                inline_cell_has_is = False
                inline_cell_has_v = False
            elif inline_cell_depth is not None and depth == inline_cell_depth + 1:
                if name == "is":
                    inline_cell_has_is = True
                elif name == "v":
                    inline_cell_has_v = True
        else:
            name = _local_name(elem.tag)
            if name == "c" and inline_cell_depth == depth:
                if inline_cell_has_v and not inline_cell_has_is:
                    return True
                inline_cell_depth = None
            elem.clear()
            depth -= 1
    return False


def _has_malformed_inline_string_value(path):
    try:
        with zipfile.ZipFile(path) as package:
            for info in package.infolist():
                name = info.filename.replace("\\", "/").lstrip("/")
                if not name.endswith(".xml") or "/worksheets/" not in f"/{name}":
                    continue
                try:
                    with package.open(info) as stream:
                        if _worksheet_has_inline_string_v_value(stream):
                            return True
                except ET.ParseError:
                    continue
    except Exception:
        return False
    return False


def _worksheet_has_empty_style_index(stream):
    for _, elem in ET.iterparse(stream, events=("start",)):
        if _local_name(elem.tag) == "c" and elem.attrib.get("s") == "":
            return True
        elem.clear()
    return False


def _has_empty_style_index(path):
    try:
        with zipfile.ZipFile(path) as package:
            names = {info.filename for info in package.infolist()}
            if "[Content_Types].xml" not in names:
                return False
            for info in package.infolist():
                name = info.filename.replace("\\", "/").lstrip("/")
                if not name.endswith(".xml") or "/worksheets/" not in f"/{name}":
                    continue
                try:
                    with package.open(info) as stream:
                        if _worksheet_has_empty_style_index(stream):
                            return True
                except ET.ParseError:
                    continue
    except Exception:
        return False
    return False


def _shared_strings_path_variant(path):
    try:
        with zipfile.ZipFile(path) as package:
            # ZipInfo.filename normalizes backslashes on Windows; orig_filename
            # preserves the package entry name so classification is host-neutral.
            names = [info.orig_filename for info in package.infolist()]
    except Exception:
        return None

    if "[Content_Types].xml" not in names:
        return None

    canonical = "xl/sharedStrings.xml"
    if canonical in names:
        return None
    for name in names:
        normalized = name.replace("\\", "/").lstrip("/")
        if normalized.lower() == "xl/sharedstrings.xml":
            return name
    return None


def _package_join(base, target):
    if target.startswith("/"):
        return posixpath.normpath(target).lstrip("/")
    return posixpath.normpath(posixpath.join(base, target)).lstrip("/")


def _styles_path_variant(path):
    try:
        with zipfile.ZipFile(path) as package:
            name_by_normalized = {
                info.filename.replace("\\", "/").lstrip("/"): info.filename
                for info in package.infolist()
            }
            names = set(name_by_normalized)
            if "[Content_Types].xml" not in names or "xl/styles.xml" in names:
                return None

            rels_candidates = [
                ("xl/workbook.xml", "xl/_rels/workbook.xml.rels"),
                ("workbook.xml", "_rels/workbook.xml.rels"),
            ]
            for workbook_path, rels_path in rels_candidates:
                if workbook_path not in names or rels_path not in names:
                    continue
                base = posixpath.dirname(workbook_path)
                with package.open(name_by_normalized[rels_path]) as stream:
                    root = ET.parse(stream).getroot()
                for rel in root.iter():
                    if _local_name(rel.tag) != "Relationship":
                        continue
                    rel_type = rel.attrib.get("Type", "")
                    if not rel_type.endswith("/styles"):
                        continue
                    target = rel.attrib.get("Target")
                    if not target:
                        continue
                    normalized = _package_join(base, target)
                    if normalized != "xl/styles.xml" and normalized in names:
                        return name_by_normalized[normalized]
    except (OSError, zipfile.BadZipFile, KeyError, ET.ParseError):
        return None
    return None


def _shared_string_lengths(package):
    try:
        with package.open("xl/sharedStrings.xml") as stream:
            root = ET.parse(stream).getroot()
    except (KeyError, ET.ParseError):
        return None
    return [len("".join(elem.itertext())) for elem in root if _local_name(elem.tag) == "si"]


def _worksheet_shared_string_expanded_chars(stream, shared_lengths):
    total = 0
    in_shared_cell = False
    shared_index = None
    for event, elem in ET.iterparse(stream, events=("start", "end")):
        name = _local_name(elem.tag)
        if event == "start":
            if name == "c" and elem.attrib.get("t") == "s":
                in_shared_cell = True
                shared_index = None
            continue
        if name == "v" and in_shared_cell:
            shared_index = elem.text
        elif name == "c" and in_shared_cell:
            try:
                index = int((shared_index or "").strip())
            except ValueError:
                index = -1
            if 0 <= index < len(shared_lengths):
                total += shared_lengths[index]
            in_shared_cell = False
        elem.clear()
    return total


def _shared_string_expanded_chars(path):
    try:
        with zipfile.ZipFile(path) as package:
            names = {info.filename for info in package.infolist()}
            if "[Content_Types].xml" not in names or "xl/sharedStrings.xml" not in names:
                return None
            shared_lengths = _shared_string_lengths(package)
            if not shared_lengths:
                return None
            total = 0
            for info in package.infolist():
                name = info.filename.replace("\\", "/").lstrip("/")
                if not name.endswith(".xml") or "/worksheets/" not in f"/{name}":
                    continue
                with package.open(info) as stream:
                    total += _worksheet_shared_string_expanded_chars(stream, shared_lengths)
            return total
    except (OSError, zipfile.BadZipFile, ET.ParseError):
        return None


def _fnum(f):
    return str(int(f)) if (f == int(f) and abs(f) < 1e15) else repr(f)


def _strip_literals(fmt):
    """Drop quoted text, [color]/[locale] brackets, and escaped chars from a
    number-format code (matches rxls classify_string), lowercased."""
    out, it = [], iter(fmt)
    for c in it:
        if c == '"':
            for c2 in it:
                if c2 == '"':
                    break
        elif c == "[":
            for c2 in it:
                if c2 == "]":
                    break
        elif c in "\\_*":
            next(it, None)
        else:
            out.append(c.lower())
    return "".join(out)


def _is_percent(fmt):
    return "%" in _strip_literals(fmt)


def _format_timedelta(value, fmt):
    low = fmt.lower().replace("\\", "")
    total_seconds = int(round(value.total_seconds()))
    sign = "-" if total_seconds < 0 else ""
    total_seconds = abs(total_seconds)
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60

    match = re.search(r"\[(h+)\]", low)
    if match:
        return f"{sign}{hours}:{minutes:02d}:{seconds:02d}"

    match = re.search(r"\[(m+)\]", low)
    if match:
        width = len(match.group(1))
        total_minutes = total_seconds // 60
        return f"{sign}{total_minutes:0{width}d}:{seconds:02d}"

    match = re.search(r"\[(s+)\]", low)
    if match:
        width = len(match.group(1))
        return f"{sign}{total_seconds:0{width}d}"

    return str(value)


def _is_numeric_time_format(fmt):
    low = _strip_literals(fmt)
    has_elapsed_time = re.search(r"\[(h+|m+|s+)\]", fmt.lower().replace("\\", "")) is not None
    has_time = has_elapsed_time or "h" in low or "s" in low
    has_date = "y" in low or "d" in low
    return has_time and not has_date


def _format_numeric_time(value, fmt):
    return _format_timedelta(datetime.timedelta(days=value), fmt)


def _cell_text(cell):
    v = cell.value
    if v is None:
        return None
    fmt = cell.number_format or ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, datetime.timedelta):
        return _format_timedelta(v, fmt)
    if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
        # Date-vs-datetime decided by the FORMAT, matching rxls/Excel.
        low = _strip_literals(fmt)
        has_time = "h" in low or "s" in low
        has_date = "y" in low or "d" in low
        if isinstance(v, datetime.time):
            return v.strftime("%H:%M:%S")
        dt = v if isinstance(v, datetime.datetime) else datetime.datetime(v.year, v.month, v.day)
        if has_date and has_time:
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        if has_time and not has_date:
            return dt.strftime("%H:%M:%S")
        return dt.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        if _is_numeric_time_format(fmt):
            return _format_numeric_time(v, fmt)
        return _fnum(v * 100) + "%" if _is_percent(fmt) else _fnum(v)
    return str(v)


def _worksheet_cell_rows(ws, max_worksheet_cells):
    declared_cells = ws.max_row * ws.max_column
    if max_worksheet_cells is not None and declared_cells > max_worksheet_cells:
        cells = getattr(ws, "_cells", None)
        if cells is None or len(cells) > max_worksheet_cells:
            raise OversizedWorksheet(
                f"{ws.title}: declared cells {declared_cells} exceed {max_worksheet_cells}"
            )
        rows = {}
        for (row, col), cell in sorted(cells.items()):
            rows.setdefault(row, []).append((col, cell))
        return [[cell for _, cell in cols] for _, cols in sorted(rows.items())]
    return ws.iter_rows()


def openpyxl_text(path, max_worksheet_cells):
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl not installed: pip install openpyxl")

    namespace = _workbook_root_namespace(path)
    if namespace == LEGACY_BETA_SPREADSHEETML_NS:
        raise UnsupportedOpenpyxlNamespace(
            f"openpyxl silently drops cells from legacy beta SpreadsheetML namespace {namespace}"
        )
    if _has_malformed_inline_string_value(path):
        raise UnsupportedOpenpyxlInlineStringValue(
            "openpyxl silently drops malformed inline-string cells that store text in <v>"
        )
    if _has_empty_style_index(path):
        raise UnsupportedOpenpyxlEmptyStyleIndex(
            "openpyxl treats empty cell style indexes as list keys"
        )
    styles_variant = _styles_path_variant(path)
    if styles_variant:
        raise UnsupportedOpenpyxlRootStylesPathVariant(
            f"openpyxl requires xl/styles.xml but package uses {styles_variant}"
        )
    shared_strings_variant = _shared_strings_path_variant(path)
    if shared_strings_variant:
        raise UnsupportedOpenpyxlSharedStringsPathVariant(
            f"openpyxl requires xl/sharedStrings.xml but package uses {shared_strings_variant}"
        )

    # NOT read_only: openpyxl's read_only mode silently drops inline-string
    # (`t="inlineStr"`) cells, which would unfairly penalise rxls (which reads
    # them). data_only yields the cached formula results rxls also reads (rxls
    # never re-evaluates formulas).
    book = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        worksheet_titles = _worksheet_titles(path)
        parts = []
        for ws in book.worksheets:
            if worksheet_titles is not None and ws.title not in worksheet_titles:
                continue
            parts.append("# " + ws.title)
            for row in _worksheet_cell_rows(ws, max_worksheet_cells):
                out = []
                for cell in row:
                    s = _cell_text(cell)
                    if s is not None and s != "":
                        out.append(s)
                if out:
                    parts.append("\t".join(out))
        return "\n".join(parts)
    finally:
        book.close()


def norm(s):
    return re.sub(r"\s+", "", s)


def hash_exact_match(left, right):
    if len(left) != len(right):
        return False
    left_hash = hashlib.sha256(left.encode("utf-8")).digest()
    right_hash = hashlib.sha256(right.encode("utf-8")).digest()
    return left_hash == right_hash


def _one_line(s):
    return re.sub(r"\s+", " ", s).strip()


_CORPUS_FAILURE_RE = re.compile(
    r"^failure:\s+(?P<ext>\S+)\s+(?P<label>.*?)\s+"
    r"kind=(?P<kind>\S+)\s+decision=(?P<decision>\S+)\s+"
    r"evidence=(?P<evidence>\S+)\s+container=(?P<container>\S+)\s+"
    r"extension_mismatch=(?P<extension_mismatch>\S+)\s+(?P<error>.*)$"
)


def parse_corpus_report(path):
    failures = []
    with open(path, encoding="utf-8") as report:
        for line in report:
            match = _CORPUS_FAILURE_RE.match(line.rstrip("\n"))
            if not match:
                continue
            failures.append(match.groupdict())
    return failures


def _path_suffix(path):
    return str(path).replace("\\", "/")


def corpus_failure_for_path(path, corpus_failures):
    if not path or not corpus_failures:
        return None
    normalized = _path_suffix(path)
    best = None
    for failure in corpus_failures:
        label = _path_suffix(failure["label"])
        if normalized.endswith(label) and (best is None or len(label) > len(best["label"])):
            best = failure
    return best


def skip_classification(kind, reason, path=None, corpus_failures=None):
    if kind == "bounded-shared-string-expansion":
        return (
            "documented_bounded_extraction",
            "shared_string_expansion_budget_exceeded",
            None,
        )
    if kind == "oversized-comparison":
        return ("needs_bounded_oracle", "comparison_budget_exceeded", None)
    if kind == "oversized-worksheet":
        return ("needs_bounded_oracle", "worksheet_guard_exceeded", None)
    if kind != "openpyxl-unreadable":
        return ("needs_oracle_triage", "unknown_skip_kind", None)

    if reason.startswith("UnsupportedOpenpyxlInlineStringValue:"):
        return ("documented_oracle_limitation", "openpyxl_malformed_inline_string", None)
    if reason.startswith("UnsupportedOpenpyxlNamespace:"):
        return ("documented_oracle_limitation", "openpyxl_namespace_limitation", None)
    if reason.startswith("UnsupportedOpenpyxlSharedStringsPathVariant:"):
        return ("documented_oracle_limitation", "openpyxl_shared_strings_path_variant", None)
    if reason.startswith("UnsupportedOpenpyxlEmptyStyleIndex:"):
        return ("documented_oracle_limitation", "openpyxl_empty_style_index", None)
    if reason.startswith("UnsupportedOpenpyxlRootStylesPathVariant:"):
        return ("documented_oracle_limitation", "openpyxl_root_styles_path_variant", None)
    if (
        reason.startswith("TypeError: expected <class 'openpyxl.styles.fills.Fill'>")
        or reason.startswith(
            "TypeError: CellStyle.__init__() got an unexpected keyword argument"
        )
        or reason.startswith(
            "TypeError: <class 'openpyxl.styles.named_styles._NamedCellStyle'>.name"
        )
    ):
        return ("documented_oracle_limitation", "openpyxl_style_parser_limitation", None)
    if reason.startswith("TypeError: Nested.from_tree() missing 1 required positional argument"):
        return ("documented_oracle_limitation", "openpyxl_pivot_parser_limitation", None)
    if reason.startswith("BadZipFile: File is not a zip file"):
        failure = corpus_failure_for_path(path, corpus_failures)
        if failure:
            return (failure["decision"], failure["evidence"], failure["kind"])
        return ("needs_corpus_crosscheck", "openpyxl_non_zip_container", None)
    if reason.startswith("BadZipFile:"):
        return ("excluded_malformed_container", "openpyxl_bad_zip", None)
    if "File contains no valid workbook part" in reason:
        return ("excluded_malformed_container", "openpyxl_no_valid_workbook_part", None)
    if "There is no item named '[Content_Types].xml'" in reason:
        return ("excluded_malformed_container", "openpyxl_missing_content_types", None)
    if "could not read manifest" in reason or "could not read strings" in reason:
        return ("excluded_malformed_container", "openpyxl_invalid_xml", None)
    if "There is no item named 'xl/sharedStrings.xml'" in reason:
        return ("needs_oracle_triage", "openpyxl_missing_shared_strings", None)
    return ("needs_oracle_triage", "openpyxl_exception", None)


def main():
    ap = argparse.ArgumentParser()
    source = ap.add_mutually_exclusive_group(required=True)
    source.add_argument("--corpus", help="dir of .xlsx/.xlsm files")
    source.add_argument("--manifest", help="public-corpus manifest.json")
    ap.add_argument("--bin", required=True, help="rxls `extract` example binary")
    ap.add_argument("--limit", type=int, default=None, help="maximum selected files")
    ap.add_argument(
        "--max-compare-chars",
        type=int,
        default=200_000,
        help="skip exact diff comparisons above this combined normalized length",
    )
    ap.add_argument(
        "--max-hash-chars",
        type=int,
        default=5_000_000,
        help="maximum combined text length eligible for oversized exact-hash admission",
    )
    ap.add_argument(
        "--max-worksheet-cells",
        type=int,
        default=5_000_000,
        help="maximum declared or loaded worksheet cells before sparse rendering/skip guards apply",
    )
    ap.add_argument(
        "--show-worst",
        type=int,
        default=0,
        help="print this many lowest-parity comparable files",
    )
    ap.add_argument(
        "--show-skips",
        type=int,
        default=0,
        help="print this many skipped files with skip kind, decision, evidence, and reason",
    )
    ap.add_argument(
        "--corpus-report",
        default=None,
        help="optional `rxls corpus-report` output used to refine non-ZIP openpyxl skip decisions",
    )
    ap.add_argument("--min", type=float, default=0.95)
    args = ap.parse_args()

    emit_parity_provenance(
        args.manifest, oracle_reader="openpyxl", package_distribution="openpyxl"
    )

    binary = resolve_binary(args.bin)
    source_root = report_source_root(args.manifest, args.corpus)
    corpus_failures = parse_corpus_report(args.corpus_report) if args.corpus_report else []

    if args.manifest:
        files = manifest_files(args.manifest, {".xlsx", ".xlsm"}, args.limit)
        print(f"manifest: {report_path(args.manifest)}")
    else:
        files = corpus_files(args.corpus, {".xlsx", ".xlsm"}, args.limit)

    sims, comparisons, skips = [], [], []
    by_skip_decision, by_skip_evidence, by_skip_corpus_kind = {}, {}, {}
    rxls_ok, opx_failed, oversized, oversized_worksheets, hash_exact = 0, 0, 0, 0, 0
    bounded_shared_strings = 0
    for f in files:
        rt = subprocess.run([binary, f], capture_output=True).stdout.decode("utf-8", "replace")
        if rt.strip():
            rxls_ok += 1
        expanded_shared_strings = _shared_string_expanded_chars(f)
        if expanded_shared_strings is not None and expanded_shared_strings > args.max_hash_chars:
            bounded_shared_strings += 1
            skips.append(
                (
                    "bounded-shared-string-expansion",
                    f,
                    (
                        f"expanded shared-string text {expanded_shared_strings} "
                        f"exceeds {args.max_hash_chars} hash budget"
                    ),
                )
            )
            continue
        try:
            gold = openpyxl_text(f, args.max_worksheet_cells)
        except OversizedWorksheet as exc:
            oversized_worksheets += 1
            skips.append(("oversized-worksheet", f, str(exc)))
            continue
        except Exception as exc:
            opx_failed += 1
            skips.append(("openpyxl-unreadable", f, f"{type(exc).__name__}: {exc}"))
            continue
        combined_text_len = len(gold) + len(rt)
        if combined_text_len > args.max_compare_chars:
            if combined_text_len > args.max_hash_chars:
                oversized += 1
                skips.append(
                    (
                        "oversized-comparison",
                        f,
                        f"combined text length {combined_text_len} exceeds {args.max_hash_chars} hash budget",
                    )
                )
                continue
            gold_norm = norm(gold)
            rxls_norm = norm(rt)
            if gold_norm and hash_exact_match(gold_norm, rxls_norm):
                hash_exact += 1
                sims.append(1.0)
                comparisons.append((1.0, f))
                continue
            oversized += 1
            skips.append(
                (
                    "oversized-comparison",
                    f,
                    f"combined text length {combined_text_len} exceeds {args.max_compare_chars}",
                )
            )
            continue
        gold_norm = norm(gold)
        rxls_norm = norm(rt)
        combined_norm_len = len(gold_norm) + len(rxls_norm)
        if combined_norm_len > args.max_compare_chars:
            if combined_norm_len > args.max_hash_chars:
                oversized += 1
                skips.append(
                    (
                        "oversized-comparison",
                        f,
                        f"combined normalized text length {combined_norm_len} exceeds {args.max_hash_chars} hash budget",
                    )
                )
                continue
            if gold_norm and hash_exact_match(gold_norm, rxls_norm):
                hash_exact += 1
                sims.append(1.0)
                comparisons.append((1.0, f))
                continue
            oversized += 1
            skips.append(
                (
                    "oversized-comparison",
                    f,
                    (
                        "combined normalized text length "
                        f"{combined_norm_len} exceeds {args.max_compare_chars}"
                    ),
                )
            )
            continue
        if not gold_norm:
            continue
        ratio = difflib.SequenceMatcher(None, gold_norm, rxls_norm).ratio()
        sims.append(ratio)
        comparisons.append((ratio, f))

    if not sims:
        sys.exit("no comparable files after oracle/size filtering")
    mean = sum(sims) / len(sims)
    print(
        f"files: {len(files)}   rxls extracted: {rxls_ok}   "
        f"openpyxl-unreadable: {opx_failed}   comparable: {len(sims)}   "
        f"oversized-comparisons: {oversized}   oversized-worksheets: {oversized_worksheets}   "
        f"hash-exact-comparisons: {hash_exact}   "
        f"bounded-shared-string-expansions: {bounded_shared_strings}"
    )
    for kind, path, reason in skips:
        decision, evidence, corpus_kind = skip_classification(
            kind, reason, path=path, corpus_failures=corpus_failures
        )
        by_skip_decision[decision] = by_skip_decision.get(decision, 0) + 1
        by_skip_evidence[evidence] = by_skip_evidence.get(evidence, 0) + 1
        if corpus_kind:
            by_skip_corpus_kind[corpus_kind] = by_skip_corpus_kind.get(corpus_kind, 0) + 1

    print(f"rxls vs openpyxl: mean parity {mean*100:.3f}%   >=99%: {sum(v>=0.99 for v in sims)}/{len(sims)}")
    for decision, count in sorted(by_skip_decision.items()):
        print(f"by_skip_decision: {decision} skipped={count}")
    for evidence, count in sorted(by_skip_evidence.items()):
        print(f"by_skip_evidence: {evidence} skipped={count}")
    for corpus_kind, count in sorted(by_skip_corpus_kind.items()):
        print(f"by_skip_corpus_kind: {corpus_kind} skipped={count}")
    for ratio, path in sorted(comparisons)[: max(0, args.show_worst)]:
        print(f"low-parity: {ratio:.3f} {report_path(path, source_root)}")
    for kind, path, reason in skips[: max(0, args.show_skips)]:
        decision, evidence, corpus_kind = skip_classification(
            kind, reason, path=path, corpus_failures=corpus_failures
        )
        corpus_part = f" corpus_kind={corpus_kind}" if corpus_kind else ""
        print(
            f"skip: kind={kind} decision={decision} evidence={evidence}{corpus_part} "
            f"path={report_path(path, source_root)} "
            f"reason={_one_line(report_reason(reason, path, source_root))}"
        )
    print(f"(note: {opx_failed} files openpyxl could not read were extracted by rxls - robustness edge)")
    sys.exit(0 if mean >= args.min else 1)


if __name__ == "__main__":
    main()
